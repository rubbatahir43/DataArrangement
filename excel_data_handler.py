from multiprocessing.reduction import duplicate
from shlex import join
from openpyxl import load_workbook, workbook
import pandas as pd
import os
import shutil
class ExcelDataHandler:

    def __init__(self, excel_path):
        self.excel_path = excel_path
        excel_new_path = self.excel_path

    def load_workbook(self, excel_new_path):
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            active_sheet_name = wb.active.title
            df = pd.read_excel(self.excel_path, sheet_name=active_sheet_name)
            return df

        return None

    def filter_excel_files(self, dispute_id, dispute_no_path):
        file_name = os.path.basename(self.excel_path)
        copied_excel_path =  os.path.join(dispute_no_path, file_name)
        shutil.copy(self.excel_path, copied_excel_path)
        
        df = self.load_workbook(copied_excel_path)
        filtered_column = dispute_id
        rows_to_hide = df[df["Dispute No."] != filtered_column].index
        work_book = load_workbook(self.excel_path)
        sheet = work_book.active
        for row_index in rows_to_hide:
            sheet.row_dimensions[row_index + 2].hidden = True

        file_name = os.path.basename(self.excel_path)
        
        file_path = os.path.join(dispute_no_path, file_name)
        work_book.save(file_path)
        visible_rows = 0
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if not sheet.row_dimensions[i].hidden:
                if any(cell.value is not None for cell in row):
                    visible_rows += 1
                    
        return visible_rows

    
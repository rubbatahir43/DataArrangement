from datetime import datetime
import os
import shutil
import re
from openpyxl import Workbook, load_workbook
import pandas as pd
from pandas._libs import missing

class PlaceData:
    def __init__(self, source_path, patient_name, service_date, 
                 dest_path, facility, dispute_no_path, claim_no, dispute_id):
        self.source_path = source_path
        self.patient_name = patient_name
        self.service_date = service_date
        self.dest_path = dest_path
        self.facility = facility
        self.dispute_no_path = dispute_no_path
        self.claim_no = claim_no
        self.dispute_id = dispute_id
        


    def convert_date_format(self, date_input, input_format, output_format):
        if isinstance(date_input, (datetime, pd.Timestamp)):
            return date_input.strftime(output_format)
        else:
            date_obj = datetime.strptime(date_input, input_format)
            return date_obj.strftime(output_format)


    def make_date_folders(self, service_date):
        dt = datetime.strptime(service_date.strip(), "%m/%d/%Y")
        year_folder = str(dt.year)
        month_folder = dt.strftime("%b %Y").upper()
        day_folder = dt.strftime("%m.%d.%Y")
        return year_folder, month_folder, day_folder

    def clean_string(self, folder_name):
        return ''.join(c for c in folder_name if c.isdigit())

    def find_month_folder(self, year_folder, month_number):
        """Find the correct month folder by matching possible month name variations."""
        month_names = {
            1: ["JAN", "JANUARY"],
            2: ["FEB", "FEBRUARY"],
            3: ["MAR", "MARCH"],
            4: ["APR", "APRIL"],
            5: ["MAY"],
            6: ["JUN", "JUNE"],
            7: ["JUL", "JULY"],
            8: ["AUG", "AUGUST"],
            9: ["SEP", "SEPT", "SEPTEMBER"],
            10: ["OCT", "OCTOBER"],
            11: ["NOV", "NOVEMBER"],
            12: ["DEC", "DECEMBER"]
        }
        possible_months = month_names[month_number]

        for folder_name in os.listdir(year_folder):
            folder_upper = folder_name.upper()
            if any(month in folder_upper for month in possible_months):
                return os.path.join(year_folder, folder_name)
        return None

    def find_date_folder(self, month_folder, service_date):
        date_obj = None
        """Find the correct date folder inside the month folder."""
        if isinstance(service_date, (datetime, pd.Timestamp)):
            date_obj = service_date
        else:
            try:
                date_obj = datetime.strptime(service_date, "%m/%d/%y")
            except ValueError:
                date_obj = datetime.strptime(service_date, "%m/%d/%Y")
                print(f"Service date format error: {service_date}")
                return None

      
        for folder_name in os.listdir(month_folder):
            if (self.compare_folder_excel(date_obj, folder_name)):
                return os.path.join(month_folder, folder_name)

        return None
    def compare_folder_excel(self, date_obj, folder_name):
        expected_full = f"{date_obj.month:02}{date_obj.day:02}{date_obj.year}"
        expected_short = f"{date_obj.month:02}{date_obj.day:02}{str(date_obj.year)[-2:]}"

        full_format, short_format  = self.clean_folder_name(folder_name)
        
        if (full_format == expected_full or
        short_format == expected_short):
            return True
        return False
        

    def clean_folder_name(self, folder_name):
        parts = re.findall(r"\d+", folder_name)

        if len(parts) == 3:
            month = parts[0].zfill(2)  # Pad single digit month
            day = parts[1].zfill(2)    # Pad single digit day
            year = parts[2]

            # Return both full and short year formats
            full_format = f"{month}{day}{year}"
            short_format = f"{month}{day}{year[-2:]}"
            return full_format, short_format

    def find_patient_folder(self, source_folder, service_date):
       
        if isinstance(service_date, (datetime, pd.Timestamp)):
            date_obj = service_date
            
        else:
            try:
                date_obj = datetime.strptime(service_date, "%m/%d/%y")
            except ValueError:
                date_obj = datetime.strptime(service_date, "%m/%d/%Y")
        print (date_obj)
        
        year_folder = os.path.join(source_folder, str(date_obj.year))
                
        if not os.path.exists(year_folder):
            print(f"Year folder not found: Could not find the specified year {year_folder}")
            return None

        # Step 1: Find Month Folder
        month_folder = self.find_month_folder(year_folder, date_obj.month)
        print (month_folder)
        
        if not month_folder:
            print("Month folder not found. Could not find the specified month")
            return None

        # Step 2: Find Date Folder
        date_folder = self.find_date_folder(month_folder, service_date)
        
        if not date_folder:
            print("Date folder not found.")
            return None

        return date_folder

    def split_patient_names(self, patient_name):
        parts = patient_name.strip().split()
        count = len(parts)

        first = parts[0] if count >=1 else ""
        last = parts[1] if count >=2 else ""
        middles = parts[1:-1]
        if count == 5:
            middle1 , middle2, middle3 = middles
            return first, middle1, middle2, middle3, last

        elif count == 4:
            middle1, middle2 = middles
            return first, middle1, middle2, last
        elif count == 3:
            middle1 = middles[0]
            return first, middle1, last
        elif count == 2:
            return first, last


    def search_folders(self, facility, patient_name, service_date):
        data_placed = False
        facility_path = os.path.join(self.source_path, facility)
        if os.path.exists(facility_path):
            
            date_folder = self.find_patient_folder(facility_path, service_date)
            
            if date_folder is not None:
                date_folder = os.path.normpath(date_folder)
            
                if os.path.isdir(date_folder):
                
                    parts = self.split_patient_names(patient_name)
                
                    for folders in os.listdir(date_folder):
                        if ("INCOMPLETE" not in folders.upper()):
                            if all(part.upper() in folders.upper() for part in parts):
                                if os.path.isdir(os.path.join(date_folder, folders)):
                                    shutil.copytree(os.path.join(date_folder, folders), os.path.join(self.dispute_no_path, folders)
                                                , dirs_exist_ok=True)

                                    data_placed = True
                        
                        
                                return True
                        elif("INCOMPLETE" in folders.upper()):
                            if all(part.upper() in folders.upper() for part in parts):
                                self.enter_Missing_data("INCOMPLETE CHARTS")

                                return False
        if not data_placed:
            self.enter_Missing_data("Patient Data Not Found")
        return False

    def enter_Missing_data(self, remarks):
        missing_data_excel = os.path.join(self.dest_path, "Missing Data.xlsx")
        column_headers = ["Patient Name", "Service Date", "Claim No", "Dispute_id", "Remarks"]
       
        self.create_workbook(missing_data_excel, column_headers, remarks)
            

    def create_workbook(self, missing_data_excel, column_headers, remarks):
        if not os.path.exists(missing_data_excel):
            wb = Workbook()
            ws = wb.active
            ws.append(column_headers)
            wb.save(missing_data_excel)

        self.load_workbook(missing_data_excel,remarks)

    def load_workbook(self, missing_data_excel, remarks):
        wb = load_workbook(missing_data_excel)
        ws = wb.active
        ws.append([self.patient_name, self.service_date, self.claim_no, self.dispute_id, remarks])
        wb.save(missing_data_excel)
        